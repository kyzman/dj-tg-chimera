import logging
import pprint

from aiogram import types, Bot
from aiogram.fsm.context import FSMContext

from openpyxl import Workbook
from openpyxl.styles import Border, Font

from dj_config.settings import XLSX_DIR
from tgbot.settings import double, thin

logger = logging.getLogger(__name__)


async def check_cart_actuality(msg: types.Message, state: FSMContext) -> bool:
    if await state.get_state():
        data = await state.get_data()
        if data.get('msg_cart_id') == msg.message_id:
            return True
    await msg.edit_reply_markup(reply_markup=None)
    await msg.answer('Выберете /cart для получения актуальной корзины')
    return False




async def clear_stored_messages(bot: Bot, current_chat_id, data: dict):
    if ids := data.get('identity'):
        if ids.get('chat_id') == current_chat_id:
            await bot.edit_message_caption(chat_id=ids.get('chat_id'), message_id=ids.get('msg_id'),
                                           caption='Вы отменили выбор', reply_markup=None)
        else:
            return
    if cart_id := data.get('msg_cart_id'):
        await bot.edit_message_text('Вы закрыли корзину', chat_id=current_chat_id, message_id=cart_id,
                                    reply_markup=None)
    else:
        return


def create_xlsx_order(data, order_no, date, address):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Заказ'
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws['A1'] = f"Заказ № {order_no:08} от {date.strftime('%d-%m-%Y %H:%M:%S')}"
    ws['A2'] = 'Артикул'
    ws['A2'].border = Border(bottom=double)
    ws['B2'] = 'Название'
    ws['B2'].border = Border(bottom=double)
    ws['C2'] = 'Цена'
    ws['C2'].border = Border(bottom=double)
    ws['D2'] = 'Количество'
    ws['D2'].border = Border(bottom=double)
    ws['E2'] = 'Сумма'
    ws['E2'].border = Border(bottom=double)
    pos = 2
    for item in data:
        pos += 1
        ws[f'A{pos}'] = item.get('pk')
        ws[f'A{pos}'].border = Border(bottom=thin)
        ws[f'B{pos}'] = item.get('name')
        ws[f'B{pos}'].border = Border(bottom=thin)
        ws[f'C{pos}'] = item.get('price')
        ws[f'C{pos}'].border = Border(bottom=thin)
        ws[f'D{pos}'] = item.get('qty')
        ws[f'D{pos}'].border = Border(bottom=thin)
        ws[f'E{pos}'] = item.get('price') * item.get('qty')
        ws[f'E{pos}'].border = Border(bottom=thin)
    pos += 1
    ws[f'A{pos}'] = f'Итоговая сумма заказа:'
    ws[f'E{pos}'] = f'=SUM(E3:E{pos - 1})'
    ws[f'E{pos}'].font = Font(bold=True)
    pos += 1
    ws[f'A{pos}'] = f'Доставка по адресу: {address}'
    try:
        wb.save(f'{XLSX_DIR}order_{order_no:08}.xlsx')
    except Exception as e:
        logger.error(e)
